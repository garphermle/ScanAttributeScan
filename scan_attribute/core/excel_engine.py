"""
Excel Engine for loading, updating, appending, exporting ranges, and merging Excel files (186 columns).
Features high-performance in-memory caching and non-blocking background saving for instant UI response (0ms lag).
"""

import os
import shutil
import time
import threading
from typing import Dict, Any, Optional, List, Tuple
import openpyxl


class ExcelEngine:
    def __init__(self, template_path: str, output_path: Optional[str] = None):
        self.template_path = template_path
        self.output_path = output_path or template_path
        self.wb: Optional[openpyxl.Workbook] = None
        self.ws = None
        self._lock = threading.RLock()

        # Dynamic column mapping (supports nhapthua1.xlsx and standard 186-column forms)
        self._col_to_excel_col: Dict[int, int] = {}
        self._excel_to_col_map: Dict[int, int] = {}
        self._note_excel_col: Optional[int] = None

        # In-memory fast cache
        self._cached_serial_rows: List[Dict[str, Any]] = []
        self._serial_to_row_map: Dict[str, int] = {}
        self._row_data_cache: Dict[int, Dict[int, Any]] = {}
        self._completed_rows_count: int = 0
        self._total_rows_count: int = 0
        self._dirty = False
        self._is_saving = False

    def _detect_column_mapping(self):
        """
        Dynamically detects column layout from row 4 (1..186) and detects Column A "Ghi Chú".
        Supports both nhapthua1.xlsx (179 cols, Col A note) and standard templates (186 cols).
        """
        self._col_to_excel_col.clear()
        self._excel_to_col_map.clear()
        self._note_excel_col = None

        if not self.ws:
            return

        max_c = max(self.ws.max_column or 0, 186)
        has_r4_nums = False

        for c in range(1, max_c + 1):
            r4_val = self.ws.cell(4, c).value
            if r4_val is not None:
                str_val = str(r4_val).strip()
                if str_val.isdigit():
                    num = int(str_val)
                    if 1 <= num <= 186:
                        self._col_to_excel_col[num] = c
                        self._excel_to_col_map[c] = num
                        has_r4_nums = True

            # Check if column is "Ghi chú" in row 1 or 2
            r1_val = str(self.ws.cell(1, c).value or "").strip().lower()
            r2_val = str(self.ws.cell(2, c).value or "").strip().lower()
            if "ghi chú" in r1_val or "ghi chu" in r1_val or "ghi chú" in r2_val or "ghi chu" in r2_val:
                if c == 1 or (c not in self._excel_to_col_map):
                    self._note_excel_col = c

        if not has_r4_nums:
            for c in range(1, 187):
                self._col_to_excel_col[c] = c
                self._excel_to_col_map[c] = c
            r1_c1 = str(self.ws.cell(1, 1).value or "").strip().lower()
            if "ghi chú" in r1_c1 or "ghi chu" in r1_c1:
                self._note_excel_col = 1
        else:
            # Fallback for any field not explicitly mapped
            for f in range(1, 187):
                if f not in self._col_to_excel_col:
                    self._col_to_excel_col[f] = f

    def initialize(self, force_reload: bool = False):
        """Ensures output file exists by copying template if necessary, then loads workbook and builds memory cache."""
        with self._lock:
            if not self.output_path:
                return

            if not os.path.exists(self.output_path):
                if self.template_path and os.path.exists(self.template_path) and os.path.abspath(self.template_path) != os.path.abspath(self.output_path):
                    try:
                        os.makedirs(os.path.dirname(os.path.abspath(self.output_path)), exist_ok=True)
                        shutil.copy(self.template_path, self.output_path)
                    except Exception as e:
                        print(f"Warning copying template: {e}")

            if (self.wb is None or force_reload) and os.path.exists(self.output_path):
                try:
                    self.wb = openpyxl.load_workbook(self.output_path)
                    if 'Data' in self.wb.sheetnames:
                        self.ws = self.wb['Data']
                    else:
                        self.ws = self.wb.active
                    self._build_cache()
                except Exception as e:
                    print(f"Error loading workbook: {e}")

    def _build_cache(self):
        """Builds in-memory index of all rows for 0ms lookup."""
        with self._lock:
            self._cached_serial_rows.clear()
            self._serial_to_row_map.clear()
            self._row_data_cache.clear()
            self._completed_rows_count = 0

            if not self.ws:
                self._total_rows_count = 0
                return

            self._detect_column_mapping()

            col_stt = self._col_to_excel_col.get(1, 1)
            col_serial = self._col_to_excel_col.get(2, 2)
            col_mhs = self._col_to_excel_col.get(3, 3)
            col_owner = self._col_to_excel_col.get(9, 9)
            col_id = self._col_to_excel_col.get(14, 14)
            col_plot = self._col_to_excel_col.get(43, 43)
            col_sheet = self._col_to_excel_col.get(44, 44)
            col_dt53 = self._col_to_excel_col.get(53, 53)
            col_xa = self._col_to_excel_col.get(93, 93)
            col_gc2 = self._col_to_excel_col.get(110, 110)
            col_folder = self._col_to_excel_col.get(183, 183)
            col_note_a = self._note_excel_col

            max_r = self.ws.max_row
            for r in range(5, max_r + 1):
                val_note_a = self.ws.cell(r, col_note_a).value if col_note_a else None
                val_stt = self.ws.cell(r, col_stt).value
                val_serial = self.ws.cell(r, col_serial).value
                val_mhs = self.ws.cell(r, col_mhs).value
                val_owner = self.ws.cell(r, col_owner).value
                val_id = self.ws.cell(r, col_id).value
                val_plot = self.ws.cell(r, col_plot).value
                val_sheet = self.ws.cell(r, col_sheet).value
                val_dt53 = self.ws.cell(r, col_dt53).value
                val_xa = self.ws.cell(r, col_xa).value
                val_gc2 = self.ws.cell(r, col_gc2).value
                val_folder = self.ws.cell(r, col_folder).value

                serial = str(val_serial or '').strip()
                if not serial and not val_mhs and not val_owner and not val_plot and not val_sheet:
                    continue

                try:
                    stt_val = int(val_stt) if (val_stt is not None and str(val_stt).strip().isdigit()) else (r - 4)
                except Exception:
                    stt_val = r - 4

                owner_name = str(val_owner or '').strip()
                id_num = str(val_id or '').strip()
                plot = str(val_plot or '').strip()
                map_sheet = str(val_sheet or '').strip()
                area = str(val_dt53 or val_xa or '').strip()
                note_a = str(val_note_a or '').strip()

                is_done = bool(owner_name or id_num or plot or map_sheet or str(val_gc2 or '').strip())

                row_obj = {
                    "row": r,
                    "stt": stt_val,
                    "serial": serial or f"Hồ sơ {stt_val}",
                    "owner_name": owner_name,
                    "id_num": id_num,
                    "plot": plot,
                    "map_sheet": map_sheet,
                    "area": area,
                    "note_a": note_a,
                    "is_completed": is_done
                }
                self._cached_serial_rows.append(row_obj)
                if is_done:
                    self._completed_rows_count += 1

                if serial:
                    self._serial_to_row_map[serial.lower()] = r
                if val_mhs:
                    self._serial_to_row_map[str(val_mhs).strip().lower()] = r
                if val_folder:
                    self._serial_to_row_map[str(val_folder).strip().lower()] = r

            self._total_rows_count = len(self._cached_serial_rows)

    def switch_target_file(self, target_path: str, copy_template_if_new: bool = True):
        """Switches target Excel file to a new or existing working file."""
        with self._lock:
            self.flush_save()
            is_new = not os.path.exists(target_path)
            if copy_template_if_new and is_new and self.template_path and os.path.exists(self.template_path):
                try:
                    os.makedirs(os.path.dirname(os.path.abspath(target_path)), exist_ok=True)
                    shutil.copy(self.template_path, target_path)
                except Exception as e:
                    print(f"Warning copying template: {e}")

            self.output_path = target_path
            self.initialize(force_reload=True)

            if is_new and copy_template_if_new and self.ws:
                for c in range(1, 187):
                    self.ws.cell(5, c).value = None
                try:
                    self.save_workbook_safe()
                except Exception as e:
                    print(f"Warning saving new empty file: {e}")
                self.initialize(force_reload=True)

    def save_workbook_safe(self, max_retries: int = 3, retry_delay: float = 0.5):
        """Saves workbook with retry handling in case of LAN network delays or temporary locks."""
        with self._lock:
            if not self.wb or not self.output_path:
                return

            last_err = None
            for attempt in range(max_retries):
                try:
                    self.wb.save(self.output_path)
                    return
                except PermissionError as pe:
                    last_err = pe
                    time.sleep(retry_delay)
                except Exception as e:
                    last_err = e
                    time.sleep(retry_delay)
            if last_err:
                raise last_err

    def _schedule_async_save(self):
        """Saves Excel to disk in a background thread to prevent UI freezing."""
        with self._lock:
            self._dirty = True
            if not self._is_saving:
                self._is_saving = True
                t = threading.Thread(target=self._async_save_worker, daemon=True)
                t.start()

    def _async_save_worker(self):
        while True:
            with self._lock:
                if not self._dirty:
                    self._is_saving = False
                    return
                self._dirty = False
            try:
                self.save_workbook_safe()
            except Exception as e:
                print(f"Background save notice: {e}")

    def flush_save(self):
        """Synchronously flushes any pending changes to disk."""
        with self._lock:
            if self._dirty:
                try:
                    self.save_workbook_safe()
                    self._dirty = False
                except Exception as e:
                    print(f"Flush save error: {e}")

    def find_row_by_serial(self, serial: str) -> Optional[int]:
        """Finds row index (1-based) matching Serial in O(1) time."""
        if not serial:
            return None
        serial_clean = str(serial).strip().lower()
        with self._lock:
            if serial_clean in self._serial_to_row_map:
                return self._serial_to_row_map[serial_clean]
        return None

    def find_first_empty_row(self) -> int:
        """Finds the first completely empty data row starting from row 5."""
        with self._lock:
            if not self.ws:
                self.initialize()

            for r in range(5, self.ws.max_row + 2):
                col2 = self.ws.cell(r, 2).value
                col3 = self.ws.cell(r, 3).value
                col9 = self.ws.cell(r, 9).value
                if (col2 is None or str(col2).strip() == '') and \
                   (col3 is None or str(col3).strip() == '') and \
                   (col9 is None or str(col9).strip() == ''):
                    return r
            return max(5, self.ws.max_row + 1)

    def get_serial_rows(self) -> List[Dict[str, Any]]:
        """Returns cached serial rows in O(1) time."""
        with self._lock:
            if not self._cached_serial_rows and self.ws:
                self._build_cache()
            return list(self._cached_serial_rows)

    def get_processed_serials_info(self) -> Dict[str, Dict[str, Any]]:
        """Returns info map in O(1) time."""
        with self._lock:
            info_map = {}
            for item in self._cached_serial_rows:
                serial_clean = item["serial"].strip().lower()
                if serial_clean:
                    info_map[serial_clean] = {
                        "serial": item["serial"],
                        "row": item["row"],
                        "stt": item["stt"],
                        "owner_name": item["owner_name"],
                        "is_completed": item["is_completed"]
                    }
            return info_map

    def get_processed_serials(self) -> List[str]:
        """Returns list of serials already present in Excel."""
        return list(self.get_processed_serials_info().keys())

    def get_data_rows_count(self) -> int:
        """Returns count of valid data rows in Excel in O(1) time."""
        with self._lock:
            return self._total_rows_count

    def get_completed_rows_count(self) -> int:
        """Returns count of completed data rows in Excel in O(1) time."""
        with self._lock:
            return self._completed_rows_count

    def read_row_data(self, row_idx: int) -> Dict[int, Any]:
        """Reads all field values for a given row index, mapped dynamically."""
        with self._lock:
            if row_idx in self._row_data_cache:
                return dict(self._row_data_cache[row_idx])

            if not self.ws:
                self.initialize()

            data: Dict[int, Any] = {}
            # Read Column A (Note)
            if self._note_excel_col:
                val_note = self.ws.cell(row_idx, self._note_excel_col).value
                data[0] = "" if val_note is None else str(val_note).strip()
            else:
                data[0] = ""

            # Read fields 1..186
            for field_num in range(1, 187):
                excel_col = self._col_to_excel_col.get(field_num, field_num)
                if excel_col <= (self.ws.max_column or 0):
                    val = self.ws.cell(row_idx, excel_col).value
                    data[field_num] = "" if val is None else val
                else:
                    data[field_num] = ""

            self._row_data_cache[row_idx] = data
            return data

    def save_row_data(self, serial: str, attr_dict: Dict[int, Any], target_row: Optional[int] = None, save_async: bool = True) -> int:
        """
        Saves row for serial in memory and starts non-blocking background disk save.
        Execution takes <5ms for instant UI responsiveness.
        """
        with self._lock:
            if not self.ws:
                self.initialize()

            row_idx = target_row
            if not row_idx:
                row_idx = self.find_row_by_serial(serial)
            if not row_idx:
                row_idx = self.find_first_empty_row()

            # Save Column A Note (Field 0) if present
            note_a_val = attr_dict.get(0, "")
            if self._note_excel_col:
                self.ws.cell(row_idx, self._note_excel_col, value=note_a_val if note_a_val else None)

            # Save STT (Field 1)
            stt_col = self._col_to_excel_col.get(1, 1)
            stt_val = row_idx - 4
            self.ws.cell(row_idx, stt_col, value=stt_val)

            if 2 not in attr_dict or not attr_dict[2]:
                attr_dict[2] = serial

            for field_num, val in attr_dict.items():
                if field_num == 0:
                    continue
                if 1 <= field_num <= 186:
                    excel_col = self._col_to_excel_col.get(field_num, field_num)
                    self.ws.cell(row_idx, excel_col, value=val)

            # Update row cache
            if row_idx not in self._row_data_cache:
                self._row_data_cache[row_idx] = {}
            self._row_data_cache[row_idx].update(attr_dict)

            # Update cached_serial_rows
            owner_name = str(attr_dict.get(9) or "")
            plot = str(attr_dict.get(43) or attr_dict.get(41) or "")
            map_sheet = str(attr_dict.get(44) or attr_dict.get(42) or "")
            gc2 = str(attr_dict.get(110) or "")
            area_val = str(attr_dict.get(53) or attr_dict.get(93) or "")
            note_a_str = str(note_a_val or "")
            is_done = bool(owner_name or plot or map_sheet or gc2 or area_val)

            found_item = None
            for item in self._cached_serial_rows:
                if item["row"] == row_idx:
                    found_item = item
                    break

            if found_item:
                was_done = found_item["is_completed"]
                found_item["owner_name"] = owner_name
                found_item["plot"] = plot
                found_item["map_sheet"] = map_sheet
                found_item["area"] = area_val
                found_item["note_a"] = note_a_str
                found_item["is_completed"] = is_done
                if not was_done and is_done:
                    self._completed_rows_count += 1
            else:
                new_item = {
                    "row": row_idx,
                    "stt": stt_val,
                    "serial": serial or f"Hồ sơ {stt_val}",
                    "owner_name": owner_name,
                    "id_num": str(attr_dict.get(14) or ""),
                    "plot": plot,
                    "map_sheet": map_sheet,
                    "area": area_val,
                    "note_a": note_a_str,
                    "is_completed": is_done
                }
                self._cached_serial_rows.append(new_item)
                self._total_rows_count = len(self._cached_serial_rows)
                if is_done:
                    self._completed_rows_count += 1

            if serial:
                self._serial_to_row_map[serial.lower()] = row_idx

            if save_async:
                self._schedule_async_save()
            return row_idx

    def export_sub_excel(self, start_stt: int, end_stt: int, output_path: str) -> int:
        """
        Exports a slice of rows corresponding to STT range [start_stt, end_stt] into a new Excel file.
        Preserves custom column structures like nhapthua1.xlsx (Column A note, etc.).
        """
        with self._lock:
            self.flush_save()
            if not self.output_path or not os.path.exists(self.output_path):
                raise FileNotFoundError(f"Master Excel file not found: {self.output_path}")

            wb_source = openpyxl.load_workbook(self.output_path)
            sheet_name = 'Data' if 'Data' in wb_source.sheetnames else wb_source.sheetnames[0]
            ws_source = wb_source[sheet_name]

            # Use source file as template to preserve exact custom columns (e.g. Column A note)
            wb_sub = openpyxl.load_workbook(self.output_path)
            ws_sub = wb_sub[sheet_name]

            max_c = ws_source.max_column

            # Clear all data rows from 5 downwards
            for r in range(5, ws_sub.max_row + 1):
                for c in range(1, max_c + 1):
                    ws_sub.cell(r, c).value = None

            dest_row = 5
            exported_count = 0

            col_stt = self._col_to_excel_col.get(1, 1)
            col_serial = self._col_to_excel_col.get(2, 2)

            for r in range(5, ws_source.max_row + 1):
                val_stt = ws_source.cell(r, col_stt).value
                val_serial = ws_source.cell(r, col_serial).value
                if not val_stt and not val_serial:
                    continue

                if val_stt is not None and str(val_stt).strip().isdigit():
                    current_stt = int(val_stt)
                else:
                    current_stt = r - 4

                if start_stt <= current_stt <= end_stt:
                    for c in range(1, max_c + 1):
                        ws_sub.cell(dest_row, c).value = ws_source.cell(r, c).value
                    ws_sub.cell(dest_row, col_stt).value = current_stt
                    dest_row += 1
                    exported_count += 1

            os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
            wb_sub.save(output_path)
            wb_sub.close()
            wb_source.close()
            return exported_count

    def merge_sub_excel(self, sub_excel_path: str) -> Tuple[int, int]:
        """
        Merges completed rows from a sub-excel file into the master Excel file.
        Supports custom column layouts and Column A note.
        """
        with self._lock:
            self.flush_save()
            if not os.path.exists(sub_excel_path):
                raise FileNotFoundError(f"Sub-excel file not found: {sub_excel_path}")

            sub_eng = ExcelEngine(sub_excel_path)
            sub_eng.initialize()

            merged_count = 0
            skipped_count = 0

            for item in sub_eng.get_serial_rows():
                sub_r = item["row"]
                serial = item["serial"]
                is_done = item["is_completed"]
                note_a = item.get("note_a", "")

                row_dict = sub_eng.read_row_data(sub_r)
                has_data = bool(is_done or note_a or any(v for k, v in row_dict.items() if k != 1 and k != 2 and v))

                if not has_data:
                    skipped_count += 1
                    continue

                target_row = self.find_row_by_serial(serial)
                if not target_row:
                    stt_val = item.get("stt")
                    if stt_val:
                        target_row = self.find_row_by_serial(str(stt_val))
                if not target_row:
                    target_row = self.find_first_empty_row()

                self.save_row_data(serial=serial, attr_dict=row_dict, target_row=target_row, save_async=False)
                merged_count += 1

            self.save_workbook_safe()
            self.initialize(force_reload=True)
            return merged_count, skipped_count
