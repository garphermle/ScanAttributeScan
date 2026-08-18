"""
Data models and Master Data Loader for ScanAttribute.
Maps to Excel_FormMau_v5_04082026.xlsx (186 columns).
"""

from dataclasses import dataclass, field
from typing import Dict, List, Optional
import datetime
import openpyxl


def parse_date_flexible(date_str: str) -> str:
    """Supports flexible date format like 1/5/2026 or 01/05/2026."""
    if not date_str:
        return ""
    date_str = str(date_str).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d.%m.%Y"):
        try:
            dt = datetime.datetime.strptime(date_str, fmt)
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            pass
    return date_str


import unicodedata


def normalize_vn_key(s: str) -> str:
    if not s:
        return ""
    s = s.strip().lower()
    for prefix in ("phường ", "xã ", "thị trấn ", "thành phố ", "thị xã ", "huyện ", "tỉnh "):
        s = s.replace(prefix, "")
    return unicodedata.normalize('NFC', s).strip()


@dataclass
class CommuneInfo:
    code_3cap: str
    name_3cap: str
    district: str
    province: str
    code_2cap: str = ""
    name_2cap: str = ""

    @property
    def full_location(self) -> str:
        parts = [self.name_3cap, self.district, self.province]
        return ", ".join(p for p in parts if p)


@dataclass
class LandTypeInfo:
    code: str
    name: str
    note: str = ""


@dataclass
class MeasurementInfo:
    commune_name: str
    district: str
    measuring_unit: str
    completion_date: str


class MasterDataManager:
    def __init__(self):
        self.communes: List[CommuneInfo] = []
        self.communes_by_code: Dict[str, CommuneInfo] = {}
        self.land_types: List[LandTypeInfo] = []
        self.land_types_by_code: Dict[str, LandTypeInfo] = {}
        self.ethnicities: List[str] = ["Kinh"]
        self.nationalities: List[str] = ["Việt Nam"]
        self.id_types: List[tuple] = [
            ("CCCD", "CCCD - Căn cước công dân"),
            ("CMND", "CMND - Chứng minh nhân dân"),
            ("HC", "HC - Hộ chiếu"),
            ("GKS", "GKS - Giấy khai sinh"),
            ("QD", "QD - Quyết định / ĐKKD"),
            ("K", "K - Khác")
        ]
        self.land_use_origins: List[tuple] = []          # List of (code, name)
        self.land_use_origins_by_code: Dict[str, str] = {}
        self.dtsd_list: List[tuple] = []                  # List of (code, name)
        self.map_types: List[tuple] = [
            ("1", "1 - Bản đồ địa chính (VN2000)"),
            ("2", "2 - Bản đồ địa chính (HN72)"),
            ("3", "3 - Bản đồ 299/TTg"),
            ("4", "4 - Sơ đồ trích đo địa chính"),
            ("5", "5 - Bản đồ địa chính khác")
        ]
        self.gcn_types: List[str] = []
        self.land_status_list: List[tuple] = []           # List of (code, name)
        self.plot_types: List[str] = []
        self.nvtc_types: List[str] = []
        self.restriction_types: List[str] = []
        self.rank_types: List[str] = []
        self.measurement_units: List[str] = []
        self.measurements: List[MeasurementInfo] = []
        self.measurements_by_key: Dict[str, MeasurementInfo] = {}

    def load_from_excel(self, excel_path: str):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            
            # Load Communes (MaXa_3cap_2cap)
            if 'MaXa_3cap_2cap' in wb.sheetnames:
                ws = wb['MaXa_3cap_2cap']
                self.communes.clear()
                self.communes_by_code.clear()
                for r in range(2, ws.max_row + 1):
                    code_3 = str(ws.cell(r, 2).value or '').strip()
                    name_3 = str(ws.cell(r, 3).value or '').strip()
                    dist = str(ws.cell(r, 4).value or '').strip()
                    prov = str(ws.cell(r, 5).value or '').strip()
                    code_2 = str(ws.cell(r, 6).value or '').strip()
                    name_2 = str(ws.cell(r, 7).value or '').strip()
                    if name_3:
                        c_info = CommuneInfo(
                            code_3cap=code_3,
                            name_3cap=name_3,
                            district=dist,
                            province=prov,
                            code_2cap=code_2,
                            name_2cap=name_2
                        )
                        self.communes.append(c_info)
                        if code_3:
                            self.communes_by_code[code_3] = c_info

            # Load Land Types (DM_LoaiDat)
            if 'DM_LoaiDat' in wb.sheetnames:
                ws = wb['DM_LoaiDat']
                self.land_types.clear()
                self.land_types_by_code.clear()
                for r in range(2, ws.max_row + 1):
                    code = str(ws.cell(r, 1).value or '').strip()
                    name = str(ws.cell(r, 2).value or '').strip()
                    note = str(ws.cell(r, 3).value or '').strip()
                    if code and name:
                        lt = LandTypeInfo(code=code, name=name, note=note)
                        self.land_types.append(lt)
                        self.land_types_by_code[code] = lt

            # Load Ethnicities
            if 'DM_DanToc' in wb.sheetnames:
                ws = wb['DM_DanToc']
                eth = [str(ws.cell(r, 2).value or '').strip() for r in range(2, ws.max_row + 1)]
                eth = [e for e in eth if e]
                other_eth = [e for e in eth if e not in ('Kinh', 'Không rõ')]
                self.ethnicities = ['Kinh', 'Không rõ'] + other_eth

            # Load Nationalities
            if 'DM_QuocTich' in wb.sheetnames:
                ws = wb['DM_QuocTich']
                nats = [str(ws.cell(r, 2).value or '').strip() for r in range(2, ws.max_row + 1)]
                nats = [n for n in nats if n]
                other_nats = [n for n in nats if n not in ('Viet Nam', 'Việt Nam', 'Vietnam')]
                self.nationalities = ['Viet Nam'] + other_nats

            # Load Land Use Origins (DM_NguonGocSuDungDat)
            if 'DM_NguonGocSuDungDat' in wb.sheetnames:
                ws = wb['DM_NguonGocSuDungDat']
                self.land_use_origins.clear()
                self.land_use_origins_by_code.clear()
                for r in range(2, ws.max_row + 1):
                    code = str(ws.cell(r, 1).value or '').strip()
                    name = str(ws.cell(r, 2).value or '').strip()
                    if code and name:
                        self.land_use_origins.append((code, name))
                        self.land_use_origins_by_code[code] = name

            # Load DTSD (DM_DoiTuongSuDungQuanLy)
            if 'DM_DoiTuongSuDungQuanLy' in wb.sheetnames:
                ws = wb['DM_DoiTuongSuDungQuanLy']
                dtsd = []
                for r in range(2, ws.max_row + 1):
                    code = str(ws.cell(r, 1).value or '').strip()
                    name = str(ws.cell(r, 2).value or '').strip().split('\n')[0].strip()
                    if code and name:
                        dtsd.append((code, name))
                if dtsd:
                    self.dtsd_list = dtsd

            # Load GCN Types (DM_LoaiGiayChungNhan)
            if 'DM_LoaiGiayChungNhan' in wb.sheetnames:
                ws = wb['DM_LoaiGiayChungNhan']
                gcns = [str(ws.cell(r, 2).value or '').strip() for r in range(2, ws.max_row + 1)]
                self.gcn_types = [g for g in gcns if g]

            # Load Land Statuses (DM_LoaiTrangThaiDangKyCapGCN)
            if 'DM_LoaiTrangThaiDangKyCapGCN' in wb.sheetnames:
                ws = wb['DM_LoaiTrangThaiDangKyCapGCN']
                statuses = []
                for r in range(2, ws.max_row + 1):
                    code = str(ws.cell(r, 1).value or '').strip()
                    name = str(ws.cell(r, 2).value or '').strip()
                    if code and name:
                        statuses.append((code, name))
                if statuses:
                    self.land_status_list = statuses

            # Load Plot Types (DM_LoaiThuaDat)
            if 'DM_LoaiThuaDat' in wb.sheetnames:
                ws = wb['DM_LoaiThuaDat']
                pts = [str(ws.cell(r, 2).value or '').strip() for r in range(2, ws.max_row + 1)]
                self.plot_types = [p for p in pts if p]

            # Load NVTC Types (DM_LoaiNghiaVuTaiChinh)
            if 'DM_LoaiNghiaVuTaiChinh' in wb.sheetnames:
                ws = wb['DM_LoaiNghiaVuTaiChinh']
                nvtcs = [str(ws.cell(r, 2).value or '').strip() for r in range(2, ws.max_row + 1)]
                self.nvtc_types = [n for n in nvtcs if n]

            # Load Restriction Types (LoaiHanChe)
            if 'LoaiHanChe' in wb.sheetnames:
                ws = wb['LoaiHanChe']
                res = [str(ws.cell(r, 2).value or '').strip() for r in range(2, ws.max_row + 1)]
                self.restriction_types = [r_val for r_val in res if r_val]

            # Load Rank Types (DM_LoaiCapHang)
            if 'DM_LoaiCapHang' in wb.sheetnames:
                ws = wb['DM_LoaiCapHang']
                ranks = [str(ws.cell(r, 2).value or '').strip() for r in range(2, ws.max_row + 1)]
                self.rank_types = [rk for rk in ranks if rk]

        except Exception as e:
            print(f"Error loading master data from {excel_path}: {e}")

        # Auto load measurement data from standard locations
        self.load_measurement_data()

    def load_measurement_data(self, meas_excel_path: str = ""):
        """Loads QNH_ThongTinDoDac.xlsx for measuring units (Col 47) and completion dates (Col 50)."""
        import os
        from scan_attribute.core.master_data import get_measurement_excel_path

        target_file = meas_excel_path if (meas_excel_path and os.path.exists(meas_excel_path)) else get_measurement_excel_path()
        if not target_file or not os.path.exists(target_file):
            return

        try:
            wb = openpyxl.load_workbook(target_file, data_only=True)
            ws = wb.active
            self.measurements.clear()
            self.measurements_by_key.clear()
            units_set = set()

            for r in range(2, ws.max_row + 1):
                c_name = str(ws.cell(r, 2).value or '').strip()
                dist = str(ws.cell(r, 3).value or '').strip()
                unit = str(ws.cell(r, 4).value or '').strip()
                date_val = ws.cell(r, 5).value

                if not c_name and not unit:
                    continue

                if isinstance(date_val, datetime.datetime):
                    date_str = date_val.strftime("%d/%m/%Y")
                else:
                    date_str = parse_date_flexible(str(date_val or ''))

                meas = MeasurementInfo(
                    commune_name=c_name,
                    district=dist,
                    measuring_unit=unit,
                    completion_date=date_str
                )
                self.measurements.append(meas)
                if unit:
                    units_set.add(unit)

                # Index by normalized keys
                k1 = f"{normalize_vn_key(c_name)}|{normalize_vn_key(dist)}"
                k2 = normalize_vn_key(c_name)
                if k1 not in self.measurements_by_key:
                    self.measurements_by_key[k1] = meas
                if k2 not in self.measurements_by_key:
                    self.measurements_by_key[k2] = meas

            self.measurement_units = sorted(list(units_set))
        except Exception as e:
            print(f"Error loading measurement data: {e}")

    def find_measurement(self, commune_name: str, district: str = "") -> Optional[MeasurementInfo]:
        """Finds measurement info matching commune and district."""
        k1 = f"{normalize_vn_key(commune_name)}|{normalize_vn_key(district)}"
        if k1 in self.measurements_by_key:
            return self.measurements_by_key[k1]
        k2 = normalize_vn_key(commune_name)
        if k2 in self.measurements_by_key:
            return self.measurements_by_key[k2]
        return None


